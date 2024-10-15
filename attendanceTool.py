import pandas as pd
from openpyxl import load_workbook

def mark_attendance(student_id, date, status, file_path, sheet):

    # Load the workbook and access the sheet to modify
    book = load_workbook(file_path)

    if sheet not in book.sheetnames:
        print(f"Sheet {sheet} not found.")
        return

    # Read the existing data of the sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name=sheet)

    # Check if the date is already a column in the DataFrame
    if date not in df.columns:
        # Add a new column for the date if it doesn't exist
        df[date] = ""

    # Find the row corresponding to the student ID
    if student_id in df['Student ID'].values:
        # Mark attendance as 'Present' or 'Absent' in the specified date column
        df.loc[df['Student ID'] == student_id, date] = status
    else:
        print(f"Student ID {student_id} not found.")
        return

    # Write the updated data back to the same sheet, preserving the other sheets
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        # No need to set writer.book manually, openpyxl handles it
        df.to_excel(writer, sheet_name=sheet, index=False)

    print(f"Attendance marked for student {student_id} on {date} as {status}")


def add_student_to_all_sheets(student_id, file_path):
    # Load the workbook
    book = load_workbook(file_path)

    # Loop through all the sheets
    for sheet_name in book.sheetnames:
        # Read the existing data of the sheet into a DataFrame
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Check if the student ID already exists in the 'Student ID' column
        if student_id not in df['Student ID'].values:
            # Append the new student ID to the DataFrame
            new_row = pd.DataFrame({'Student ID': [student_id]})
            df = pd.concat([df, new_row], ignore_index=True)

            # Write the updated DataFrame back to the same sheet
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            print(f"Student ID {student_id} added to {sheet_name}")
        else:
            print(f"Student ID {student_id} already exists in {sheet_name}")

    print("Student ID addition complete.")
